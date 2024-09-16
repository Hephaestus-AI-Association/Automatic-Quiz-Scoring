import os
import shutil
import cv2
import pdf2image
import numpy as np
import pandas as pd
from skimage.transform import resize 
from skimage.io import imread
import joblib
from openpyxl.styles import PatternFill, NamedStyle, Alignment, Font, Border, Side
from openpyxl.utils.cell import get_column_letter

# Color variables
HEADER = '\033[95m'
ENDC = '\033[0m'


# Make temp trash directory
if os.path.exists(f'{os.getcwd()}/trash'):
    shutil.rmtree(f'{os.getcwd()}/trash')
[os.makedirs(f'{os.getcwd()}/trash/{dir}') for dir in ['1_files', '2_cropped', '3_labelling_img']]


print(f'{HEADER}Preprocessing files...{ENDC}')
# Read uploaded files and get jpeg images. Accepted types: jpeg, jpg, png, PDF
def extract_images(out_folder, in_folder=f'{os.getcwd()}/predict/files'):
    """
    Takes input files and convert them to jpeg.
    Uploads output jpeg files to desired output folder.
    Pictures are numbered. Pictures extracted from a PDF are numbered both by file and page

    Parameters
    ----------
    out_folder : str
        The folder to which upload output pictures
    """

    raw_files = os.listdir(in_folder)
    file_num = 1
    for file in raw_files:
        print(f'Processing file {file}')
        ext = file.split('.')[-1]

        # jpg or jpeg
        if ext.lower() == 'jpg' or ext.lower() == 'jpeg':
            img = cv2.imread(f'{in_folder}/{file}')
            cv2.imwrite(f'{out_folder}/{file_num}.jpeg', img)

        # png
        elif ext.lower() == 'png':
            img = cv2.imread(f'{in_folder}/{file}')
            cv2.imwrite(f'{out_folder}/{file_num}.jpeg', img, [int(cv2.IMWRITE_JPEG_QUALITY), 100])

        # pdf
        elif ext.lower() == 'pdf':
            pages = pdf2image.convert_from_path(f'{in_folder}/{file}', poppler_path='poppler-24.02.0/Library/bin')
            page_num = 1
            for page in pages:
                page.save(f'{out_folder}/{file_num}_{page_num}.jpeg', 'JPEG')
                page_num += 1

        print('Done.')
        file_num += 1


print(f'{HEADER}\nExtracting images...{ENDC}')
extract_images(f'{os.getcwd()}/trash/1_files')


def crop_images(input_folder, output_folder):
    """
    Finds area of interest from images applying a classifier and crops images accordingly.
    Also, images are filtered and rotated. Images are saved in the desired folder

    Parameters
    ----------
    input_folder : str
        The folder from which import pictures
    output_folder : str
        The folder to which upload output pictures
    """

    classifier = cv2.CascadeClassifier('cascade/classifier/cascade.xml')

    for filename in os.listdir(input_folder):
        image_path = os.path.join(input_folder, filename)

        image = cv2.imread(image_path)
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

        # creates a box taking the coordinates of the furthest to the right, left, top and bottom objects
        objects = classifier.detectMultiScale(gray)
        if len(objects) == 4:
            x_min = min(objects[:, 0])
            y_min = min(objects[:, 1])
            x_max = max(objects[:, 0] + objects[:, 2])
            y_max = max(objects[:, 1] + objects[:, 3])


            # get points
            s_objects = objects[np.lexsort(np.fliplr(objects).T)]
            if s_objects[0,1] < s_objects[1,1]:
                p1 = [s_objects[0,0], s_objects[0,1]]
                p3 = [s_objects[1,0], s_objects[1,1] + s_objects[1,3]]
            else:
                p1 = [s_objects[1,0], s_objects[1,1]]
                p3 = [s_objects[0,0], s_objects[0,1] + s_objects[0,3]]

            if s_objects[2,1] < s_objects[3,1]:
                p2 = [s_objects[2,0] + s_objects[2,2], s_objects[2,1]]
                p4 = [s_objects[3,0] + s_objects[3,2], s_objects[3,1] + s_objects[3,3]]
            else:
                p2 = [s_objects[3,0] + s_objects[3,2], s_objects[3,1]]
                p4 = [s_objects[2,0] + s_objects[2,2], s_objects[2,1] + s_objects[2,3]]

            points = np.array([p2, p4, p3, p1], np.float32)
            (tl, tr, br, bl) = points

            widthA = np.sqrt(((br[0] - bl[0]) ** 2) + ((br[1] - bl[1]) ** 2))
            widthB = np.sqrt(((tr[0] - tl[0]) ** 2) + ((tr[1] - tl[1]) ** 2))
            maxWidth = max(int(widthA), int(widthB))

            heightA = np.sqrt(((tr[0] - br[0]) ** 2) + ((tr[1] - br[1]) ** 2))
            heightB = np.sqrt(((tl[0] - bl[0]) ** 2) + ((tl[1] - bl[1]) ** 2))
            maxHeight = max(int(heightA), int(heightB))

            dst = np.array([
                [0, 0],
                [maxWidth - 1, 0],
                [maxWidth - 1, maxHeight - 1],
                [0, maxHeight - 1]], dtype = "float32")

            # compute the perspective transform matrix and then apply it
            M = cv2.getPerspectiveTransform(points, dst)
            warped = cv2.warpPerspective(gray, M, (maxWidth, maxHeight))

            # detects rotated files and adjust them
            if image.shape[1]>image.shape[0]:
                if x_min < image.shape[1] - x_max:
                    warped = cv2.rotate(warped, cv2.ROTATE_90_COUNTERCLOCKWISE)
                else:
                    warped = cv2.rotate(warped, cv2.ROTATE_90_CLOCKWISE)
                print(' rotated ↴')

            warped = cv2.rotate(warped, cv2.ROTATE_90_CLOCKWISE)
            warped = cv2.resize(warped, (1400, 1870)) 

            cropped_filename = os.path.join(output_folder, f"cropped_{filename}")

            cv2.imwrite(cropped_filename, warped)
            print(f"Saved cropped image: {cropped_filename}")


print(f'{HEADER}\nCropping images...{ENDC}')
crop_images(f'{os.getcwd()}/trash/1_files', f'{os.getcwd()}/trash/2_cropped')


def apply_mask(image_path):
    """
    Mask is applied to image to extract table entries

    Parameters
    ----------
    image_path : str
        The picture you want to apply mask to

    Returns
    -------
    list
        a list of coordinates of the boxes
    """

    img = cv2.imread(image_path)
    img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    mask = cv2.imread('mask.png',0)
    res = cv2.bitwise_and(img, img, mask = mask)

    connected_components = cv2.connectedComponentsWithStats(res, 4, cv2.CV_32S)

    boxes = get_boxes(connected_components)
    return boxes


def get_boxes(connected_components):
    """
    Obtains box coordinates

    Parameters
    ----------
    connected_components : Array
        cv2.connectedComponentsWithStats Object

    Returns
    -------
    list
        a list of coordinates of the boxes
    """

    (totalLabels, label_ids, values, centroid) = connected_components

    slots = []
    coef = 1

    for i in range(1, totalLabels):

        x1 = int(values[i, cv2.CC_STAT_LEFT] * coef)
        y1 = int(values[i, cv2.CC_STAT_TOP] * coef)
        w = int(values[i, cv2.CC_STAT_WIDTH] * coef)
        h = int(values[i, cv2.CC_STAT_HEIGHT] * coef)

        slots.append([x1,y1,w,h])

    return slots


def classify(input_folder, output_folder):
    """
    Stores each box in a separate file

    Parameters
    ----------
    input_folder : str
        The folder from which import pictures
    output_folder : str
        The folder to which upload output pictures
    """

    for filename in os.listdir(input_folder):
        image_path = os.path.join(input_folder, filename)

        boxes = apply_mask(image_path)

        image = cv2.imread(image_path)
        image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

        boxes_sorted = sort_boxes(boxes)

        if boxes_sorted.shape[0] != 50:
            continue

        count = 0
        for box in boxes_sorted:

            cv2.imwrite(f'{output_folder}/' + filename.replace('.jpeg', f'_{count}.jpeg'), image[box[1]:box[1]+box[3], box[0]:box[0]+box[2]])

            count += 1


def sort_boxes(boxes):
    """
    Sorts boxes according to their coordinates

    Parameters
    ----------
    boxes : list
        list of boxes' coordinates

    Returns
    -------
    list
        a list of sorted boxes
    """

    boxes_df = pd.DataFrame(boxes)
    return boxes_df.sort_values([0,1]).to_numpy()


print(f'{HEADER}\nExtracting answers...{ENDC}')
classify(f'{os.getcwd()}/trash/2_cropped', f'{os.getcwd()}/trash/3_labelling_img')


PCA = joblib.load('PCA_model.pkl')
model = joblib.load('SVC_model.pkl')
predictions = pd.DataFrame(np.zeros(shape=(50,5), dtype=np.int16), index=(list(range(1,51))), columns=['A','B','C','D','0'])

# Split elements in chunks of size 1000
# Import images into DataFrame Object and predicting labels
print(f'\n{HEADER}Processing images and predicting...{ENDC}')

l = os.listdir(f'{os.getcwd()}/trash/3_labelling_img')
chunks = np.array_split(l, len(l)//1000)
count = 0
for chunk in chunks:
    print(f'{count / len(chunks) * 100:.1f}% done.')

    print(' | Creating DataFrame')
    flat_data_arr=[]
    for x in chunk:
        img_array = imread(f'{os.getcwd()}/trash/3_labelling_img/{x}') 
        img_resized = resize(img_array,(96,130,3))
        flat_data_arr.append(img_resized.flatten())

    print(' | Predicting')
    df = pd.DataFrame(flat_data_arr)

    y_pred = model.predict(PCA.transform(df))

    for i in range(df.shape[0]):
        q = int(chunk[i].split('_')[-1].split('.')[0])
        predictions.loc[int(q)+1, y_pred[i]] += 1
    
    count += 1

print(f'100% done.')


# Excel export and styling
print(f'{HEADER}\nExporting to Excel...{ENDC}')
predictions.to_excel("predict/output_report.xlsx")  
with pd.ExcelWriter("predict/output_report.xlsx", engine="openpyxl") as writer:
    predictions.to_excel(writer, sheet_name='Report')

    workbook = writer.book
    worksheet = workbook["Report"]

    worksheet.insert_rows(1)
    worksheet.cell(1, 1).value = "Automatic Quiz Scoring Report"
    worksheet.cell(1, 1).font = Font(bold=True, size=14)
    worksheet.cell(1, 1).alignment = Alignment(horizontal="center")
    worksheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)


shutil.rmtree(f'{os.getcwd()}/trash')


print(f'{HEADER}\nDone!{ENDC}')
